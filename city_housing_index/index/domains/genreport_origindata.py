import openpyxl
from openpyxl.styles import Side, Border, Font, Alignment

class GenOriginalReport:
    def __init__(self,url) -> None:
        self.edge = Side(style='thin', color='000000')
        self.border = Border(top=self.edge, bottom=self.edge, left=self.edge, right=self.edge)
        self.url = url
        self.center = Alignment(horizontal='center')
        self.report = openpyxl.Workbook()
    def create_report(self):
        totalmarket_vol = self.report.active
        totalmarket_vol.title = '整体市场-样本量'
        totalmarket_index = self.report.create_sheet()
        totalmarket_index.title = '整体市场-原始指数'
        area1_vol = self.report.create_sheet()
        area1_vol.title = '面积1-样本量'
        area1_index = self.report.create_sheet()
        area1_index.title = '面积1-原始指数'
        area2_vol = self.report.create_sheet()
        area2_vol.title = '面积2-样本量'
        area2_index = self.report.create_sheet()
        area2_index.title = '面积2-原始指数'
        area3_vol = self.report.create_sheet()
        area3_vol.title = '面积3-样本量'
        area3_index = self.report.create_sheet()
        area3_index.title = '面积3-原始指数'
        check = self.report.create_sheet()
        check.title = '当月核对'
        return

    def get_date(self, t:int) -> str:
        year = t // 12 + 2006
        month = t % 12 + 1
        return str(year) + '年' + str(month) + '月'
    
    def to_char(self, t:int) -> str:
        if t < 26:
            return chr(0x41 + t)
        return self.to_char(t // 26) + self.to_char(t % 26)


    def Market(self, cityList):
        # 每个city中应有参数
        # {'code', 'name', 'trade_vol':list, 'index':list, 'trade_vol_under_90':list, 'index_under_90':list,
        #  'trade_vol_90_144':list, 'index_90_144':list, 'trade_vol_above_144':list, 'index_above_144':list}
        table = {}
        key_list = ['totalmarket_vol','totalmarket_index','area1_vol','area1_index','area2_vol', 'area2_index','area3_vol','area3_index']
        table_name = ['整体市场-样本量','整体市场-原始指数', '面积1-样本量', '面积1-原始指数', '面积2-样本量', '面积2-原始指数', '面积3-样本量', '面积3-原始指数']
        for i in range(0,len(key_list)):
            table[key_list[i]] = self.report.worksheets[i]

        # 表格格式设置
        datafont = Font(u'Times New Roman', size=10, bold=False, color='000000')
        for key in key_list:
            table[key].freeze_panes = 'C2'
            table[key].column_dimensions['A'].width = 5.0
        col_max = len(cityList[0]['index']) + 3

        # 填充时间
        for key in key_list:
            for i in range(3, col_max):
                cell = table[key].cell(row=1, column=i)
                cell.value = self.get_date(i - 3)
                cell.alignment = self.center
                cell.font = datafont

        # 填充城市数据
        now_row = 2
        for city in cityList:
            # 填充城市代码与城市名
            for key in key_list:
                cell = table[key].cell(row=now_row, column=1)
                cell.value = city['code']
                cell.alignment = self.center
                cell.font = datafont
                cell = table[key].cell(row=now_row, column=2)
                cell.value = city['name']
                cell.alignment = self.center
                cell.font = datafont
            # 填充整数型数据
            int_key = ['totalmarket_vol', 'area1_vol', 'area2_vol', 'area3_vol']
            int_data = ['trade_vol', 'trade_vol_under_90', 'trade_vol_90_144', 'trade_vol_above_144']
            for i in range(0,len(int_key)):
                for j in range(3, col_max):
                    cell = table[int_key[i]].cell(row=now_row, column=j)
                    if city[int_data[i]][j - 3] > 0:
                        cell.value = city[int_data[i]][j - 3]
                    else:
                        cell.value = ''
                    cell.alignment = self.center
                    cell.font = datafont
            # 填充浮点数型数据
            float_key = ['totalmarket_index', 'area1_index', 'area2_index', 'area3_index']
            float_data = ['index', 'index_under_90', 'index_90_144', 'index_above_144']
            for i in range(0,len(float_key)):
                for j in range(3, col_max):
                    cell = table[float_key[i]].cell(row=now_row, column=j)
                    if city[int_data[i]][j - 3] > 0:
                        cell.value = float('%.2f' % city[float_data[i]][j - 3])
                    else:
                        cell.value = ''
                    cell.alignment = self.center
                    cell.font = datafont
            now_row += 1
        
        # 添加排序功能
        for key in key_list:
            table[key].auto_filter.ref = "A1:" + self.to_char(col_max - 1) + str(len(cityList) + 1)
            table[key].auto_filter.add_sort_condition("A1:A" + str(len(cityList) + 1))

    def Check(self, cityList):
        # 每个city中应有参数
        # {'code', 'name', 'chain', 'volumn', 'chain_under_90', 'vol_under_90',
        #  'chain_90_144', 'vol_90_144', 'chain_above_144', 'vol_above_144', 'max_area', 'max_price'}
        check = self.report['当月核对']
        datafont = Font(u'Times New Roman', size=10, bold=False, color='000000')
        redfont = Font(u'Times New Roman', size=10, bold=False, color='ff0000')
        check.freeze_panes = 'C2'
        check.column_dimensions['A'].width = 5.0
        col_max = 15
        title = ['整体市场环比', '面积1环比', '面积2环比', '面积3环比', '整体市场同比', '面积1同比', '面积2同比', '面积3同比', '样本量差值', '最大面积', '最高单价', '情况备注']
        for i in range(0,len(title)):
            cell = check.cell(row=1, column=3 + i)
            cell.value = title[i]
            cell.font = datafont
            cell.alignment = self.center
        now_row = 2
        for city in cityList:
            # 填入整数型数据与字符串
            int_date = ['code', 'name', 'volumn', 'vol_under_90', 'vol_90_144', 'vol_above_144', 'max_price']
            float_data = ['chain', 'chain_under_90', 'chain_90_144', 'chain_above_144', 'max_area']
            int_col = [1, 2, 7, 8, 9, 10, 13]
            float_col = [3, 4, 5, 6, 12]
            for i in range(0,len(int_date)):
                cell = check.cell(row=now_row, column=int_col[i])
                if city[int_date[i]] == 0:
                    cell.value = ''
                else:
                    cell.value = city[int_date[i]]
                cell.font = datafont
                cell.alignment = self.center
            for i in range(0,len(float_data)):
                cell = check.cell(row=now_row, column=float_col[i])
                if city[float_data[i]] == 0:
                    cell.value = ''
                    cell.font = datafont
                elif city[float_data[i]]<0:
                    cell.value = float('%.2f' % (-city[float_data[i]]*100))
                    cell.font = redfont
                else:
                    cell.value = float('%.2f' % (city[float_data[i]]*100))
                    cell.font = datafont
                cell.alignment = self.center
            now_row += 1
        check.auto_filter.ref = "A1:" + self.to_char(col_max - 1) + str(len(cityList) + 1)
        check.auto_filter.add_sort_condition("A1:A" + str(len(cityList) + 1))

    def EndReport(self):
        self.report.save(self.url)