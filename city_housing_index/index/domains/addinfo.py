from django.contrib import admin
from index.models import CalculateResult
from django.core.exceptions import ObjectDoesNotExist
from index.models import CityIndex
from city.models import City
from city.enums import CityArea
from calculate.models import DataFile
import openpyxl


def get_data_info(year: int, month: int, city: int):
    datafile_list = DataFile.objects.filter(city_code=city)
    url = ''
    exist = False
    for datafile in datafile_list:
        time = datafile.start.split('-')
        if len(time) < 2:
            continue
        if int(time[0]) == year and int(time[1]) == month:
            url = 'media/' + str(datafile.file)
            exist = True
            # break
    if not exist:
        return 0
    data = openpyxl.load_workbook(url)
    sheet_list = data.get_sheet_names()
    sheet = data.get_sheet_by_name(sheet_list[0])
    num_col = sheet.max_column
    num_row = sheet.max_row
    col_name = []
    total_area = 0
    total_price = 0

    total_area_under_90 = 0
    total_area_90_144 = 0
    total_area_above_144 = 0

    total_price_under_90 = 0
    total_price_90_144 = 0
    total_price_above_144 = 0

    volume_under_90 = 0
    volume_90_144 = 0
    volume_above_144 = 0

    max_area = 0
    max_price = 0

    fromdata = {}
    for i in range(1, num_col + 1):
        col_name.append(sheet.cell(row=1, column=i).value.upper())
        fromdata[col_name[i - 1]] = []
        for j in range(2, num_row + 1):
            fromdata[col_name[i - 1]].append(sheet.cell(row=j, column=i).value)
    for i in range(0, num_row - 1):
        area = float(fromdata['UNIT_AREA'][i])
        if area > max_area:
            max_area = area
        price = float(fromdata['UNIT_PRICE'][i])
        if price > max_price:
            max_price = price
        price = area * price
        total_area += area
        total_price += price
        if area < 0.9:
            total_area_under_90 += area
            total_price_under_90 += price
            volume_under_90 += 1
        elif area < 1.44:
            total_area_90_144 += area
            total_price_90_144 += price
            volume_90_144 += 1
        else:
            total_area_above_144 += area
            total_price_above_144 += price
            volume_above_144 += 1
    try:
        price = total_price / total_area
    except ZeroDivisionError:
        price = 0
    try:
        price_under_90 = total_price_under_90 / total_area_under_90
    except ZeroDivisionError:
        price_under_90 = 0
    try:
        price_90_144 = total_price_90_144 / total_area_90_144
    except ZeroDivisionError:
        price_90_144 = 0
    try:
        price_above_144 = total_price_above_144 / total_area_above_144
    except ZeroDivisionError:
        price_above_144 = 0

    re = {
        'trade_volume': num_row - 1,
        'volume_under_90': volume_under_90,
        'volume_90_144': volume_90_144,
        'volume_above_144': volume_above_144,
        'price': price,
        'area_volume': total_area,
        'area_volume_under_90': total_area_under_90,
        'area_volume_above_144': total_area_above_144,
        'area_volume_90_144': total_area_90_144,
        'price_under_90': price_under_90,
        'price_above_144': price_above_144,
        'price_90_144': price_90_144,
        'max_area': max_area,
        'max_price': max_price
    }
    return re


def add_new_month(year: int, month: int):
    # add city part
    city_list = City.objects.filter(ifin90=True)
    for city in city_list:
        try:
            newline = CalculateResult.objects.get(year=year, month=month, city_or_area=True, city=int(city.code))
        except ObjectDoesNotExist:
            newline = CalculateResult(year=year, month=month, city_or_area=True, city=int(city.code))
        newline.save()
    for i in range(0, CityArea.num_of_item_90):
        try:
            newline = CalculateResult.objects.get(year=year, month=month, city_or_area=False, area=i)
        except ObjectDoesNotExist:
            newline = CalculateResult(year=year, month=month, city_or_area=False, area=i)
        newline.save()
    return True


def upload_city_info_to_database(year, month, city):
    re = get_data_info(year, month, city)
    if re == 0:
        return False
    try:
        row = CityIndex.objects.get(year=year, month=month, city=city)
        row.delete()
    except ObjectDoesNotExist:
        pass
    newline = CityIndex(year=year,
                        month=month,
                        city=city,
                        price=re['price'],
                        trade_volume=re['trade_volume'],
                        area_volume=re['area_volume'],
                        area_volume_under_90=re['area_volume_under_90'],
                        area_volume_above_144=re['area_volume_above_144'],
                        area_volume_90_144=re['area_volume_90_144'],
                        price_under_90=re['price_under_90'],
                        price_above_144=re['price_above_144'],
                        price_90_144=re['price_90_144'],
                        trade_volume_under_90=re['volume_under_90'],
                        trade_volume_above_144=re['volume_above_144'],
                        trade_volume_90_144=re['volume_90_144'],
                        max_area=re['max_area'],
                        max_price=re['max_price'])
    newline.save()
    return True
