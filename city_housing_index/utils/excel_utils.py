import xlrd
import openpyxl


def read_xls(file_path):
    workbook = xlrd.open_workbook(file_path, encoding_override="utf_8")
    sheet = workbook.sheet_by_index(0)
    nrows = sheet.nrows
    ncols = sheet.ncols
    title = []
    content = []

    for i in range(0, ncols):
        title.append(sheet.cell(0, i).value)

    for i in range(1, nrows):
        tmp = []
        for j in range(0, ncols):
            tmp.append(sheet.cell(i, j).value)
        content.append(tmp)
    return title, content


def read_xlsx(file_path):
    workbook = openpyxl.open(file_path)
    sheet_name = workbook.sheetnames[0]
    sheet = workbook.get_sheet_by_name(sheet_name)
    rows = sheet.max_row
    columns = sheet.max_column
    title = []
    content = []

    for i in range(1, columns + 1):
        title.append(sheet.cell(1, i).value)

    for i in range(2, rows + 1):
        tmp = []
        for j in range(1, columns + 1):
            tmp.append(sheet.cell(i, j).value)
        content.append(tmp)
    return title, content


def write_xlsx_contact(file_name, title, content):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "通讯录"
    for i in range(1, len(title) + 1):
        sheet.cell(1, i).value = title[i - 1]

    for row in range(1, len(content) + 1):
        for col in range(1, len(content[row]) + 1):
            sheet.cell(row=row, column=col).value = content[row - 1][col - 1]
    workbook.save(filename=file_name)

def write_xlsx(file_name, sheet_name, title, content):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(1, len(title) + 1):
        sheet.cell(1, i).value = title[i - 1]

    for row in range(1, len(content) + 1):
        for col in range(1, len(content[row]) + 1):
            sheet.cell(row=row, column=col).value = content[row - 1][col - 1]
    workbook.save(filename=file_name)
