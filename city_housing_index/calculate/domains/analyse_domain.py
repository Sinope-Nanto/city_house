from sklearn import linear_model
import openpyxl
import math
import time

def get_model_parameter(block_list:list, data:list[dict]):
    num = len(data)
    switch = num / 10 # 是否设置哑变量的阈值

    pro_id_list = {}
    for i in range(0, num):
        if not data[i]['PRO_ID'] in pro_id_list.keys():
            pro_id_list[data[i]['PRO_ID']] = {'list':[i], 'num': 1, 'mean_price':data[i]['UNIT_PRICE']}
        else:
            pro_id_list[data[i]['PRO_ID']]['list'].append(i)
            pro_id_list[data[i]['PRO_ID']]['mean_price'] = (pro_id_list[data[i]['PRO_ID']]['mean_price'] * pro_id_list[data[i]['PRO_ID']]['num'] + data[i]['UNIT_PRICE']) \
                / (pro_id_list[data[i]['PRO_ID']]['num'] + 1)
            pro_id_list[data[i]['PRO_ID']]['num'] += 1
    dummy_var = []
    dummy_mean_var = {}
    for pro_id in pro_id_list.keys():
        if pro_id_list[pro_id]['num'] >= switch and pro_id != '' and pro_id_list[pro_id]['num'] < num:
            dummy_var.append('dummy' + pro_id)
            dummy_mean_var['dummy' + pro_id] = pro_id_list[pro_id]['num'] / num
            for row in data:
                row['dummy' + pro_id] = 0
            for id in pro_id_list[pro_id]['list']:
                data[id]['dummy' + pro_id] = 1
    
    block_var = []
    for i in range(1, len(block_list)):
        block_var.append(block_list[i])
    
    liner_var = ['PRO_AREA', 'PRO_FLOOR', 'UNIT_DURATION', 'UNIT_FLOOR', 'UNIT_ONSALE', 'ZX_CU', 'ZX_JING']
    quadratic_var = ['PRO_AREA*PRO_AREA', 'PRO_FLOOR*PRO_FLOOR', 'UNIT_DURATION*UNIT_DURATION', 'UNIT_FLOOR*UNIT_FLOOR', 'UNIT_AREA*UNIT_AREA']
    for i in range(0, 4):
        for j in range(0,num):
            data[j][quadratic_var[i]] = data[j][liner_var[i]]**2
    for j in range(0,num):
        data[j]['UNIT_AREA*UNIT_AREA'] = data[j]['UNIT_AREA']
    var_list = block_var + liner_var + quadratic_var + dummy_var 
    var_mean_value = {}
    for key in block_var + liner_var + quadratic_var:
        var_mean_value[key] = 0
    dataset = []
    y = []
    geometric_mean_value = 0
    for i in range(0,num):
        for key in block_var + liner_var + quadratic_var:
            try:
                var_mean_value[key] += data[i][key]
            except KeyError:
                pass
        sample = []
        geometric_mean_value += math.log(data[i]['UNIT_PRICE'])
        y.append(math.log(data[i]['UNIT_PRICE']))
        for var in var_list:
            try:
                sample.append(data[i][var])
            except KeyError:
                sample.append(0)
        dataset.append(sample)
    for key in block_var + liner_var + quadratic_var:
        var_mean_value[key] /= num
    geometric_mean_value = math.exp(geometric_mean_value/num)
    reg = linear_model.LinearRegression()
    reg.fit(X=dataset,y=y)
    for pro_id in pro_id_list.keys():
        pro_id_list[pro_id]['predict_price'] = math.exp(reg.predict([dataset[pro_id_list[pro_id]['list'][0]]])[0])
        pro_id_list[pro_id]['radio'] = pro_id_list[pro_id]['mean_price']/pro_id_list[pro_id]['predict_price'] - 1
    main_var_factors = 0
    for i in range(len(block_var + liner_var + quadratic_var), len(var_list)):
        main_var_factors += (dummy_mean_var[var_list[i]] * reg.coef_[i]) 
    model_data = {
        'intercept': reg.intercept_,
        'coef' : {},
        'var_mean_value' : var_mean_value,
        'geometric_mean_value' : geometric_mean_value,
        'main_var_factors' : main_var_factors,
        'num_of_data' : num
    }
    for i in range(0, len(block_var + liner_var + quadratic_var)): 
        model_data['coef'][var_list[i]] = reg.coef_[i] 
    return pro_id_list, model_data

def load_data(url):
    data = openpyxl.load_workbook(url)
    sheet = data[data.sheetnames[0]]
    num_col = sheet.max_column
    num_row = sheet.max_row
    col_name = []
    fromdata = []
    for i in range(1, num_col + 1):
        col_name.append(sheet.cell(row=1, column=i).value.upper())
    for i in range(2, num_row + 1):
        fromdata.append({})
        for j in range(1, num_col + 1):
            if col_name[j - 1] == 'PRO_ID' or col_name[j - 1] == 'DEAL_TIME':
                    fromdata[i - 2][col_name[j - 1]] = str(sheet.cell(row=i, column=j).value)
            else:
                    fromdata[i - 2][col_name[j - 1]] = float(sheet.cell(row=i, column=j).value)
    return fromdata

def gen_html_report(data_url, last_data_url, save_url, block_list, last_block_list, template_url, year, month):
    template = open(template_url, 'r', encoding='utf-8').read()
    now = int(round(time.time()*1000))
    datatime = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(now/1000))
    template = template.replace('$$NUMBER$$','Y'+ str(year - 2000).zfill(2) + str(month).zfill(2))
    template = template.replace('$$DATA$$', datatime)
    template = template.replace('$$PERSON$$', 'XXX')
    trade_list, model = get_model_parameter(block_list, load_data(data_url))
    last_trade_list, last_model = get_model_parameter(last_block_list, load_data(last_data_url))
    geometric_mean_radio = model['geometric_mean_value'] / last_model['geometric_mean_value'] - 1
    LRmodel = 'Ln( 理论价格 ) = ' + ('%.2f' % model['intercept']) + '+ (主体变量影响)'
    for key in model['coef'].keys():
        if(model['coef'][key] >= 0):
            LRmodel += (' + ' + '%.3f' % model['coef'][key] + ' <i>' + key + '</i>')
        else:
            LRmodel += (' - ' + '%.3f' % (-model['coef'][key]) + ' <i>' + key + '</i>')
    template = template.replace('$$LRMODEL$$', LRmodel)

    table1 = ''
    for key in trade_list:
        if abs(trade_list[key]['radio']) < 0.5 or (trade_list[key]['num'] / model['num_of_data'] < 0.02):
            table1 += '<tr><td>{}</td><td>{:d}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td></tr>\n'\
                .format(key, trade_list[key]['num'], trade_list[key]['predict_price'], trade_list[key]['mean_price'], trade_list[key]['radio'])
        else:
            table1 += '<tr bgcolor=\"yellow\"><td>{}</td><td>{:d}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td></tr>\n'\
                .format(key, trade_list[key]['num'], trade_list[key]['predict_price'], trade_list[key]['mean_price'], trade_list[key]['radio'])

    table2 = ''
    structural_factors = 1
    for key in model['coef'].keys():
        coef = model['coef'][key]
        value = model['var_mean_value'][key]
        if key in last_model['coef'].keys():
            last_val = last_model['var_mean_value'][key]
        else:
            last_val = 0
        rate_of_change = (value - last_val) * coef
        structural_factors *= math.exp(rate_of_change)
        if abs(rate_of_change) < 0.02:
            table2 += '<tr><td>{}</td><td>{:.4f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td>\n'.\
                format(key, coef, last_val, value, value - last_val, rate_of_change)
        else:
            table2 += '<tr bgcolor=\"yellow\"><td>{}</td><td>{:.4f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td>\n'.\
                format(key, coef, last_val, value, value - last_val, rate_of_change)
    rate_of_change = (model['main_var_factors'] - last_model['main_var_factors'])
    if abs(rate_of_change) < 0.02:
        table2 += '<tr><td>主体变量综合效应</td><td>1.0000</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td>\n'.\
            format(last_model['main_var_factors'], model['main_var_factors'], rate_of_change, rate_of_change)
    else:
        table2 += '<tr bgcolor=\"yellow\"><td>主体变量综合效应</td><td>1.0000</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2f}</td><td>{:.2%}</td>\n'.\
            format(last_model['main_var_factors'], model['main_var_factors'], rate_of_change, rate_of_change)

    structural_factors *= math.exp(rate_of_change)
    structural_factors = structural_factors - 1
    template = template.replace('$$TABLE_1$$', table1)
    template = template.replace('$$TABLE_2$$', table2)
    template = template.replace('$$Geometric_mean$$', "{:.2%}".format(geometric_mean_radio))
    template = template.replace('$$Structural_factors$$', "{:.2%}".format(structural_factors))
    open(save_url, 'w+', encoding='utf-8').write(template)
    return save_url